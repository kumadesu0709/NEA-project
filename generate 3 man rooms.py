from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import ast
import random


wb = load_workbook("")
ws = wb.active

class Student:

    def __init__(self, name, preferred_one, preferred_two, preferred_three, hated, is_weekly):
        self._name = name
        self._preferred_one = preferred_one
        self._preferred_two = preferred_two
        self._preferred_three = preferred_three
        self._hated = hated
        if is_weekly == "Yes":
            self._is_weekly = True
        else:
            self._is_weekly = False

    def name(self):
        return self._name
    
    def preferred_one(self):
        return self._preferred_one
    
    def preferred_two(self):
        return self._preferred_two
    
    def preferred_three(self):
        return self._preferred_three
    
    def hated(self):
        return self._hated
    
    def is_weekly(self):
        return self._is_weekly
    
    def __str__(self):
        return str(self._name)

class rooming:

    def __init__(self, student_in_yeargroup:list, unwanted_pairs:list, wanted_pairs:list, weekly_boarder_setting:str, amount_of_combinations: int):
        self._student_in_yeargroup = student_in_yeargroup
        self._unwanted_pairs = unwanted_pairs
        self._wanted_pairs = wanted_pairs
        self._amount_of_combinations = amount_of_combinations
        self._rooming_combinations = []
        self._rooming_scores = {}

        self._weekly_boarders = []
        for student in student_in_yeargroup:
            if student.is_weekly() == True:
                self._weekly_boarders.append(student.name())

        self._weekly_boarders_setting = weekly_boarder_setting
        if self._weekly_boarders_setting == "Put all the weekly boarders into the same room":
            if len(self._weekly_boarders) <= 3:
                self._wanted_pairs.append([self._weekly_boarders])
        elif self._weekly_boarders_setting == "Pair each of them with a full boarder":
            for i in range (len(self._weekly_boarders)):
                for j in range (i+1, len(self._weekly_boarders)):
                    self._unwanted_pairs.append([self._weekly_boarders[i],self._weekly_boarders[j]])
    
    def generate_all_combination(self, max_no_of_people):
        result = []
        def backtrack(start, comb):
            if len(comb) == max_no_of_people:
                result.append(comb.copy())
                return
            for i in range(start, len(self._student_in_yeargroup)):
                comb.append(self._student_in_yeargroup[i])
                backtrack(i+1, comb)
                comb.pop()
        backtrack(0, [])
        return result


    def _check_existed_pupil(self,room_one, room_two):
        for student_one in room_one:
            if student_one in room_two:
                return True
        return False

    def pick_rooms(self, rooms:list, no_of_rooms:int):
        result = []
        def backtrack(start, comb):
            if len(comb) == no_of_rooms:
                result.append(comb.copy())
                return
            for i in range(start, len(rooms)):
                have_existed = False
                for existed_room in comb:
                    if self._check_existed_pupil(rooms[i], existed_room):
                        have_existed = True
                        break
                if have_existed == False:
                    comb.append(rooms[i])
                    backtrack(i+1,comb)
                    comb.pop()

        backtrack(0,[])
        return result
    
    def _produce_rooming__without_a_type_of_room(self, combs_room_one, combs_room_two):
        for comb_room_one in combs_room_one:
            for comb_room_two in combs_room_two:
                for room_one in comb_room_one:
                    for room_two in comb_room_two:
                        if self._check_existed_pupil(room_one, room_two):
                            continue
                        self._rooming_combinations.append([comb_room_one, comb_room_two])

    
    def produce_roomings(self, no_of_one_man_room:int,no_of_two_man_room:int,no_of_three_man_room:int):

        one_man_rooms = self.generate_all_combination(1)
        two_man_rooms = self.generate_all_combination(2)
        three_man_rooms = self.generate_all_combination(3)

        combs_one_man_rooms = self.pick_rooms(one_man_rooms,no_of_one_man_room)
        combs_two_man_rooms = self.pick_rooms(two_man_rooms, no_of_two_man_room)
        combs_three_man_rooms = self.pick_rooms(three_man_rooms, no_of_three_man_room)

        if no_of_two_man_room == 0 and no_of_three_man_room == 0 and no_of_one_man_room > 0:
            for comb in combs_one_man_rooms:
                self._rooming_combinations.append([comb])
        
        elif no_of_one_man_room == 0 and no_of_three_man_room == 0 and no_of_two_man_room > 0:
            for comb in combs_two_man_rooms:
                self._rooming_combinations.append([comb])
        
        elif no_of_two_man_room == 0 and no_of_one_man_room == 0 and no_of_three_man_room > 0:
            for comb in combs_three_man_rooms:
                self._rooming_combinations.append([comb])

        elif no_of_one_man_room == 0 and no_of_two_man_room > 0 and no_of_three_man_room > 0:
            self._produce_rooming__without_a_type_of_room(combs_two_man_rooms, combs_three_man_rooms)

        elif no_of_one_man_room > 0 and no_of_two_man_room == 0 and no_of_three_man_room > 0:
            self._produce_rooming__without_a_type_of_room(combs_one_man_rooms, combs_three_man_rooms)

        elif no_of_one_man_room > 0 and no_of_two_man_room > 0 and no_of_three_man_room == 0:
            self._produce_rooming__without_a_type_of_room(combs_one_man_rooms, combs_two_man_rooms)

        elif no_of_one_man_room > 0 and no_of_two_man_room > 0 and no_of_three_man_room > 0:
            for comb_one_man_room in combs_one_man_rooms:
                for comb_two_man_room in combs_two_man_rooms:
                    for one_man_room in comb_one_man_room:
                        for two_man_room in comb_two_man_room:
                            if self._check_existed_pupil(one_man_room, two_man_room):
                                continue
                            for comb_three_man_room in combs_three_man_rooms:
                                for three_man_room in comb_three_man_room:
                                    if self._check_existed_pupil(three_man_room, one_man_room) or self._check_existed_pupil(two_man_room, three_man_room):
                                        continue
                                    self._rooming_combinations.append([comb_one_man_room,comb_two_man_room,comb_three_man_room])

    
    def _randomly_pick_combination(self, amount_of_combination):
        result = []
        random_locations = random.sample(range(len(self._rooming_combinations)),amount_of_combination)
        for location in random_locations:
            result.append(self._rooming_combinations[location])
        self._rooming_combinations = result
            

    def _give_score_to_room(self,room:list):

        score = 0

        if len(room) == 2:

            if room[0].preferred_one() == room[1].name():
                score = score + 5
            if room[0].preferred_two() == room[1].name():
                score = score + 3
            if room[0].preferred_three() == room[1].name():
                score = score + 1
            if room[0].hated() == room[1]:
                score = score - 1


            if room[1].preferred_one() == room[0].name():
                score = score + 5
            if room[1].preferred_two() == room[0].name():
                score = score + 3
            if room[1].preferred_three() == room[0].name():
                score = score + 1
            if room[1].hated() == room[0].name():
                score = score - 1
        
        if len(room) == 3:

            if room[0].preferred_one() == room[1].name() or room[0].preferred_one == room[2].name():
                score = score + 5
            if room[0].preferred_two() == room[1].name() or room[0].preferred_two() == room[2].name():
                score = score + 3
            if room[0].preferred_three() == room[1].name() or room[0].preferred_three() == room[2].name():
                score = score + 1
            if room[0].hated() == room[1].name() or room[0].hated() == room[2].name():
                score = score - 1


            if room[1].preferred_one() == room[0].name() or room[1].preferred_one() == room[2].name():
                score = score + 5
            if room[1].preferred_two() == room[0].name() or room[1].preferred_two() == room[2].name():
                score = score + 3
            if room[1].preferred_three() == room[0].name() or room[1].preferred_three() == room[2].name():
                score = score + 1
            if room[1].hated() == room[0].name() or room[1].hated() == room[2].name():
                score = score - 1
            
            if room[2].preferred_one() == room[0].name() or room[2].preferred_one() == room[1].name():
                score = score + 5
            if room[2].preferred_two() == room[0].name() or room[2].preferred_two() == room[1].name():
                score = score + 3
            if room[2].preferred_three() == room[0].name() or room[2].preferred_three() == room[1].name():
                score = score + 1
            if room[2].hated() == room[0].name() or room[2].hated() == room[1].name():
                score = score - 1
            
        return score
    
    def give_score_to_combinations(self,randomly_pick_room_on:bool):
        if randomly_pick_room_on == True:
            self._randomly_pick_combination(500)
        for combination in self._rooming_combinations:
            score = 0
            rooms_in_combination = []
            for x_man_room_combination in combination:
                for room in x_man_room_combination:
                    score += self._give_score_to_room(room)
                    student_in_room = []
                    for student in room:
                        student_in_room.append(student.name())
                    rooms_in_combination.append(student_in_room)
            self._rooming_scores[str(rooms_in_combination)] = score
            self._clean()
    
    def _clean(self):

        rooming_scores_keys = self._rooming_scores.copy()
        need_to_remove = False

        if len(self._wanted_pairs) > 0 and len(self._unwanted_pairs) > 0:
            for key in rooming_scores_keys.keys():
                combination = ast.literal_eval(key)
                need_to_remove = True
                for room in combination:
                    for wanted_pair in self._wanted_pairs:
                        for unwanted_pair in self._unwanted_pairs:
                            if set(wanted_pair).issubset(set(room)) == True and set(unwanted_pair).issubset(set(room)) == False:
                                need_to_remove = False
        
        elif len(self._wanted_pairs) > 0  and len(self._unwanted_pairs) == 0:
            for key in rooming_scores_keys.keys():
                combination = ast.literal_eval(key)
                need_to_remove = True
                for room in combination:
                    for wanted_pair in self._wanted_pairs:
                        if set(wanted_pair).issubset(set(room)) == True:
                            need_to_remove = False
        
        elif len(self._unwanted_pairs) > 0 and len(self._wanted_pairs) == 0:
            for key in rooming_scores_keys.keys():
                combination = ast.literal_eval(key)
                need_to_remove = True
                for room in combination:
                    for unwanted_pair in self._unwanted_pairs:
                        if set(unwanted_pair).issubset(set(room)) == True:
                            need_to_remove = False

        if need_to_remove == True:
            if key in self._rooming_scores.keys():
                self._rooming_scores.pop(key)
        
        self._rooming_scores = dict(sorted(self._rooming_scores.items(), key=lambda item: item[1], reverse=True))
        self._rooming_scores = {k: self._rooming_scores[k] for k in list(self._rooming_scores)[:self._amount_of_combinations]}

    def return_rooming(self):
        return self._rooming_combinations
    
    def return_scores(self):
        return self._rooming_scores
    




student_one = Student(ws["B2"].value, ws["C2"].value, ws["D2"].value, ws["E2"].value, ws["F2"].value, ws["G2"].value )
student_two = Student(ws["B3"].value, ws["C3"].value, ws["D3"].value, ws["E3"].value, ws["F3"].value, ws["G3"].value)
student_three = Student(ws["B4"].value, ws["C4"].value, ws["D4"].value, ws["E4"].value, ws["F4"].value, ws["G4"].value)
student_four = Student(ws["B5"].value, ws["C5"].value, ws["D5"].value, ws["E5"].value, ws["F5"].value, ws["G5"].value)
student_five = Student(ws["B6"].value, ws["C6"].value, ws["D6"].value, ws["E6"].value, ws["F6"].value, ws["G6"].value)
student_six = Student(ws["B7"].value, ws["C7"].value, ws["D7"].value, ws["E7"].value, ws["F7"].value, ws["G7"].value)
student_seven = Student(ws["B8"].value, ws["C8"].value, ws["D8"].value, ws["E8"].value, ws["F8"].value, ws["G8"].value)
student_eight = Student(ws["B9"].value, ws["C9"].value, ws["D9"].value, ws["E9"].value, ws["F9"].value, ws["G9"].value)
student_nine = Student(ws["B10"].value, ws["C10"].value, ws["D10"].value, ws["E10"].value, ws["F10"].value, ws["G10"].value)
student_ten = Student(ws["B11"].value, ws["C11"].value, ws["D11"].value, ws["E11"].value, ws["F11"].value, ws["G11"].value)
student_eleven = Student(ws["B12"].value, ws["C12"].value, ws["D12"].value, ws["E12"].value, ws["F12"].value, ws["G12"].value)
student_twelve= Student(ws["B13"].value, ws["C13"].value, ws["D13"].value, ws["E13"].value, ws["F13"].value, ws["G13"].value)
student_thirteen = Student(ws["B14"].value, ws["C14"].value, ws["D14"].value, ws["E14"].value, ws["F14"].value, ws["G14"].value)

students = [student_one,student_two, student_three, student_four, student_five, student_six, student_seven, student_eight, student_nine, student_ten, student_eleven, student_twelve, student_thirteen]
while (students[-1]).name() == None:
    students.pop()
rooming1 = rooming(students, [], [], "a", 100)
rooming1.produce_roomings(0,5,0)
rooming1.give_score_to_combinations(True)
print(rooming1.return_scores())
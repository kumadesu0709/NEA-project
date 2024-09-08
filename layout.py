import PyQt6.QtWidgets as qtw
from PyQt6.QtGui import * 
import sys
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import ast
import random

class Student:

    def __init__(self, name, preferred_one, preferred_two, preferred_three, hated, is_weekly):
        self._name = name
        self._preferred_one = preferred_one
        self._preferred_two = preferred_two
        self._preferred_three = preferred_three
        self._hated = hated
        if is_weekly == "Yes":
            self._is_weekly = True
        else:
            self._is_weekly = False

    def name(self):
        return self._name
    
    def preferred_one(self):
        return self._preferred_one
    
    def preferred_two(self):
        return self._preferred_two
    
    def preferred_three(self):
        return self._preferred_three
    
    def hated(self):
        return self._hated
    
    def is_weekly(self):
        return self._is_weekly
    
    def __str__(self):
        return str(self._name)

class rooming:

    def __init__(self, student_in_yeargroup:list, unwanted_pairs:list, wanted_pairs:list, weekly_boarder_setting:str, amount_of_combinations: int):
        self._student_in_yeargroup = student_in_yeargroup
        self._unwanted_pairs = unwanted_pairs
        self._wanted_pairs = wanted_pairs
        self._amount_of_combinations = amount_of_combinations
        self._rooming_combinations = []
        self._rooming_scores = {}

        self._weekly_boarders = []
        for student in student_in_yeargroup:
            if student.is_weekly() == True:
                self._weekly_boarders.append(student.name())

        self._weekly_boarders_setting = weekly_boarder_setting
        if self._weekly_boarders_setting == "Put all the weekly boarders into the same room":
            if len(self._weekly_boarders) <= 3:
                self._wanted_pairs.append([self._weekly_boarders])
        elif self._weekly_boarders_setting == "Pair each of them with a full boarder":
            for i in range (len(self._weekly_boarders)):
                for j in range (i+1, len(self._weekly_boarders)):
                    self._unwanted_pairs.append([self._weekly_boarders[i],self._weekly_boarders[j]])
    
    def generate_all_combination(self, max_no_of_people):
        result = []
        def backtrack(start, comb):
            if len(comb) == max_no_of_people:
                result.append(comb.copy())
                return
            for i in range(start, len(self._student_in_yeargroup)):
                comb.append(self._student_in_yeargroup[i])
                backtrack(i+1, comb)
                comb.pop()
        backtrack(0, [])
        return result


    def _check_existed_pupil(self,room_one, room_two):
        for student_one in room_one:
            if student_one in room_two:
                return True
        return False

    def pick_rooms(self, rooms:list, no_of_rooms:int):
        result = []
        def backtrack(start, comb):
            if len(comb) == no_of_rooms:
                result.append(comb.copy())
                return
            for i in range(start, len(rooms)):
                have_existed = False
                for existed_room in comb:
                    if self._check_existed_pupil(rooms[i], existed_room):
                        have_existed = True
                        break
                if have_existed == False:
                    comb.append(rooms[i])
                    backtrack(i+1,comb)
                    comb.pop()

        backtrack(0,[])
        return result
    
    def _produce_rooming__without_a_type_of_room(self, combs_room_one, combs_room_two):
        for comb_room_one in combs_room_one:
            for comb_room_two in combs_room_two:
                for room_one in comb_room_one:
                    for room_two in comb_room_two:
                        if self._check_existed_pupil(room_one, room_two):
                            continue
                        self._rooming_combinations.append([comb_room_one, comb_room_two])

    
    def produce_roomings(self, no_of_one_man_room:int,no_of_two_man_room:int,no_of_three_man_room:int):

        one_man_rooms = self.generate_all_combination(1)
        two_man_rooms = self.generate_all_combination(2)
        three_man_rooms = self.generate_all_combination(3)

        combs_one_man_rooms = self.pick_rooms(one_man_rooms,no_of_one_man_room)
        combs_two_man_rooms = self.pick_rooms(two_man_rooms, no_of_two_man_room)
        combs_three_man_rooms = self.pick_rooms(three_man_rooms, no_of_three_man_room)

        if no_of_two_man_room == 0 and no_of_three_man_room == 0 and no_of_one_man_room > 0:
            for comb in combs_one_man_rooms:
                self._rooming_combinations.append([comb])
        
        elif no_of_one_man_room == 0 and no_of_three_man_room == 0 and no_of_two_man_room > 0:
            for comb in combs_two_man_rooms:
                self._rooming_combinations.append([comb])
        
        elif no_of_two_man_room == 0 and no_of_one_man_room == 0 and no_of_three_man_room > 0:
            for comb in combs_three_man_rooms:
                self._rooming_combinations.append([comb])

        elif no_of_one_man_room == 0 and no_of_two_man_room > 0 and no_of_three_man_room > 0:
            self._produce_rooming__without_a_type_of_room(combs_two_man_rooms, combs_three_man_rooms)

        elif no_of_one_man_room > 0 and no_of_two_man_room == 0 and no_of_three_man_room > 0:
            self._produce_rooming__without_a_type_of_room(combs_one_man_rooms, combs_three_man_rooms)

        elif no_of_one_man_room > 0 and no_of_two_man_room > 0 and no_of_three_man_room == 0:
            self._produce_rooming__without_a_type_of_room(combs_one_man_rooms, combs_two_man_rooms)

        elif no_of_one_man_room > 0 and no_of_two_man_room > 0 and no_of_three_man_room > 0:
            for comb_one_man_room in combs_one_man_rooms:
                for comb_two_man_room in combs_two_man_rooms:
                    for one_man_room in comb_one_man_room:
                        for two_man_room in comb_two_man_room:
                            if self._check_existed_pupil(one_man_room, two_man_room):
                                continue
                            for comb_three_man_room in combs_three_man_rooms:
                                for three_man_room in comb_three_man_room:
                                    if self._check_existed_pupil(three_man_room, one_man_room) or self._check_existed_pupil(two_man_room, three_man_room):
                                        continue
                                    self._rooming_combinations.append([comb_one_man_room,comb_two_man_room,comb_three_man_room])

    
    def _randomly_pick_combination(self, amount_of_combination):
        result = []
        random_locations = random.sample(range(len(self._rooming_combinations)),amount_of_combination)
        for location in random_locations:
            result.append(self._rooming_combinations[location])
        self._rooming_combinations = result
            

    def _give_score_to_room(self,room:list):

        score = 0

        if len(room) == 2:

            if room[0].preferred_one() == room[1].name():
                score = score + 5
            if room[0].preferred_two() == room[1].name():
                score = score + 3
            if room[0].preferred_three() == room[1].name():
                score = score + 1
            if room[0].hated() == room[1]:
                score = score - 1


            if room[1].preferred_one() == room[0].name():
                score = score + 5
            if room[1].preferred_two() == room[0].name():
                score = score + 3
            if room[1].preferred_three() == room[0].name():
                score = score + 1
            if room[1].hated() == room[0].name():
                score = score - 1
        
        if len(room) == 3:

            if room[0].preferred_one() == room[1].name() or room[0].preferred_one == room[2].name():
                score = score + 5
            if room[0].preferred_two() == room[1].name() or room[0].preferred_two() == room[2].name():
                score = score + 3
            if room[0].preferred_three() == room[1].name() or room[0].preferred_three() == room[2].name():
                score = score + 1
            if room[0].hated() == room[1].name() or room[0].hated() == room[2].name():
                score = score - 1


            if room[1].preferred_one() == room[0].name() or room[1].preferred_one() == room[2].name():
                score = score + 5
            if room[1].preferred_two() == room[0].name() or room[1].preferred_two() == room[2].name():
                score = score + 3
            if room[1].preferred_three() == room[0].name() or room[1].preferred_three() == room[2].name():
                score = score + 1
            if room[1].hated() == room[0].name() or room[1].hated() == room[2].name():
                score = score - 1
            
            if room[2].preferred_one() == room[0].name() or room[2].preferred_one() == room[1].name():
                score = score + 5
            if room[2].preferred_two() == room[0].name() or room[2].preferred_two() == room[1].name():
                score = score + 3
            if room[2].preferred_three() == room[0].name() or room[2].preferred_three() == room[1].name():
                score = score + 1
            if room[2].hated() == room[0].name() or room[2].hated() == room[1].name():
                score = score - 1
            
        return score
    
    def give_score_to_combinations(self,randomly_pick_room_on:bool):
        if randomly_pick_room_on == True:
            self._randomly_pick_combination(500)
        for combination in self._rooming_combinations:
            score = 0
            rooms_in_combination = []
            for x_man_room_combination in combination:
                for room in x_man_room_combination:
                    score += self._give_score_to_room(room)
                    student_in_room = []
                    for student in room:
                        student_in_room.append(student.name())
                    rooms_in_combination.append(student_in_room)
            self._rooming_scores[str(rooms_in_combination)] = score
            self._clean()
    
    def _clean(self):

        rooming_scores_keys = self._rooming_scores.copy()
        need_to_remove = False

        if len(self._wanted_pairs) > 0 and len(self._unwanted_pairs) > 0:
            for key in rooming_scores_keys.keys():
                combination = ast.literal_eval(key)
                need_to_remove = True
                for room in combination:
                    for wanted_pair in self._wanted_pairs:
                        for unwanted_pair in self._unwanted_pairs:
                            if set(wanted_pair).issubset(set(room)) == True and set(unwanted_pair).issubset(set(room)) == False:
                                need_to_remove = False
        
        elif len(self._wanted_pairs) > 0  and len(self._unwanted_pairs) == 0:
            for key in rooming_scores_keys.keys():
                combination = ast.literal_eval(key)
                need_to_remove = True
                for room in combination:
                    for wanted_pair in self._wanted_pairs:
                        if set(wanted_pair).issubset(set(room)) == True:
                            need_to_remove = False
        
        elif len(self._unwanted_pairs) > 0 and len(self._wanted_pairs) == 0:
            for key in rooming_scores_keys.keys():
                combination = ast.literal_eval(key)
                need_to_remove = True
                for room in combination:
                    for unwanted_pair in self._unwanted_pairs:
                        if set(unwanted_pair).issubset(set(room)) == True:
                            need_to_remove = False

        if need_to_remove == True:
            if key in self._rooming_scores.keys():
                self._rooming_scores.pop(key)
        
        self._rooming_scores = dict(sorted(self._rooming_scores.items(), key=lambda item: item[1], reverse=True))
        self._rooming_scores = {k: self._rooming_scores[k] for k in list(self._rooming_scores)[:self._amount_of_combinations]}

    def return_rooming(self):
        return self._rooming_combinations
    
    def return_scores(self):
        return self._rooming_scores

class MainWindow(qtw.QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Rooming Generator")
        self.setting_window = None

        layout = qtw.QGridLayout()
        vbox_layout = qtw.QVBoxLayout()

        form_layout = qtw.QFormLayout()
        self.directory = qtw.QLineEdit("")
        form_layout.addRow("Please type the directory of the excel file here (instructions below):", self.directory)
        vbox_layout.addLayout(form_layout)

        non_editable_line_edit = qtw.QLineEdit(parent=self)
        non_editable_line_edit.setReadOnly(True)
        non_editable_line_edit.setText("https://www.digitaltrends.com/computing/how-to-find-and-copy-a-file-path-on-mac/")
        vbox_layout.addWidget(non_editable_line_edit)

        self.submit_button = qtw.QPushButton("Submit")
        self.submit_button.clicked.connect(self.submit_button_clicked)
        vbox_layout.addWidget(self.submit_button)

        layout.addLayout(vbox_layout,0,0,1,3)

        widget = qtw.QWidget()
        widget.setLayout(layout)

        self.setCentralWidget(widget)
    
    def submit_button_clicked(self):
        if self.setting_window == None:

            window.close()
            wb = load_workbook(self.directory.text())
            ws = wb.active

            student_one = Student(ws["B2"].value, ws["C2"].value, ws["D2"].value, ws["E2"].value, ws["F2"].value, ws["G2"].value)
            student_two = Student(ws["B3"].value, ws["C3"].value, ws["D3"].value, ws["E3"].value, ws["F3"].value, ws["G3"].value)
            student_three = Student(ws["B4"].value, ws["C4"].value, ws["D4"].value, ws["E4"].value, ws["F4"].value, ws["G4"].value)
            student_four = Student(ws["B5"].value, ws["C5"].value, ws["D5"].value, ws["E5"].value, ws["F5"].value, ws["G5"].value)
            student_five = Student(ws["B6"].value, ws["C6"].value, ws["D6"].value, ws["E6"].value, ws["F6"].value, ws["G6"].value)
            student_six = Student(ws["B7"].value, ws["C7"].value, ws["D7"].value, ws["E7"].value, ws["F7"].value, ws["G7"].value)
            student_seven = Student(ws["B8"].value, ws["C8"].value, ws["D8"].value, ws["E8"].value, ws["F8"].value, ws["G8"].value)
            student_eight = Student(ws["B9"].value, ws["C9"].value, ws["D9"].value, ws["E9"].value, ws["F9"].value, ws["G9"].value)
            student_nine = Student(ws["B10"].value, ws["C10"].value, ws["D10"].value, ws["E10"].value, ws["F10"].value, ws["G10"].value)
            student_ten = Student(ws["B11"].value, ws["C11"].value, ws["D11"].value, ws["E11"].value, ws["F11"].value, ws["G11"].value)
            student_eleven = Student(ws["B12"].value, ws["C12"].value, ws["D12"].value, ws["E12"].value, ws["F12"].value, ws["G12"].value)
            student_twelve= Student(ws["B13"].value, ws["C13"].value, ws["D13"].value, ws["E13"].value, ws["F13"].value, ws["G13"].value)
            student_thirteen = Student(ws["B14"].value, ws["C14"].value, ws["D14"].value, ws["E14"].value, ws["F14"].value, ws["G14"].value)

            students = [student_one,student_two, student_three, student_four, student_five, student_six, student_seven, student_eight, student_nine, student_ten, student_eleven, student_twelve, student_thirteen]
            while (students[-1]).name() == None:
                students.pop()
            self.setting_window = SettingsWindow(students)
            self.setting_window.show()
        
        else:

            self.setting_window.close()
            self.setting_window = None


class SettingsWindow(qtw.QWidget):
    def __init__(self, students):

        super().__init__()
        self.setWindowTitle("Settings")

        self._unwanted_pairs = []
        self._wanted_pairs = []
        self._weekly_setting_text = ""
        
        self.vlayout = qtw.QVBoxLayout()
        self.unwanted_pairs_hbox_layout = qtw.QHBoxLayout()
        self.unwanted_pairs_vbox_layout = qtw.QVBoxLayout()
        self.wanted_pairs_hbox_layout = qtw.QHBoxLayout()
        self.wanted_pairs_vbox_layout = qtw.QVBoxLayout()
        self.unwanted_pairs_buttons = qtw.QHBoxLayout()
        self.wanted_pairs_buttons = qtw.QHBoxLayout()
        self.pairs_display = qtw.QHBoxLayout()

        self.amount_of_combination_layout = qtw.QHBoxLayout()
        self.weekly_settings = qtw.QVBoxLayout()
        self.room_settings = qtw.QGridLayout()


        self._students = students
        self.result_window = None
        self._unwanted_combination_text = ""
        self._wanted_combination_text = ""

        self.one_man_label = qtw.QLabel("One Man Rooms")
        self.amount_of_one_man = qtw.QComboBox()
        self.amount_of_one_man.addItems(str(i) for i in range(0,(len(self._students)+1)))
        self.two_man_label = qtw.QLabel("Two Man Rooms")
        self.amount_of_two_man = qtw.QComboBox()
        self.amount_of_two_man.addItems(str(i) for i in range(0,(len(self._students)//2+1)))
        self.three_man_label = qtw.QLabel("Three Man Rooms")
        self.amount_of_three_man = qtw.QComboBox()
        self.amount_of_three_man.addItems(str(i) for i in range(0,(len(self._students)//3+1)))
        self.room_settings.addWidget(self.one_man_label,0,0)
        self.room_settings.addWidget(self.two_man_label,0,1)
        self.room_settings.addWidget(self.three_man_label,0,2)
        self.room_settings.addWidget(self.amount_of_one_man,1,0)
        self.room_settings.addWidget(self.amount_of_two_man,1,1)
        self.room_settings.addWidget(self.amount_of_three_man,1,2)
        self.vlayout.addLayout(self.room_settings)




        self.amount_of_combinations_label = qtw.QLabel("How many combinations would you like to see?")
        self.amount_of_combinations = qtw.QLineEdit("")
        self.amount_of_combination_layout.addWidget(self.amount_of_combinations_label)
        self.amount_of_combination_layout.addWidget(self.amount_of_combinations)
        self.vlayout.addLayout(self.amount_of_combination_layout)


        self.weekly_boarder_label = qtw.QLabel("Would you like to...")
        self.all_in_same_room = qtw.QRadioButton("Put all the weekly boarders into the same room")
        self.all_in_same_room.clicked.connect(self.all_in_same_room_button_clicked)
        self.put_them_in_pairs = qtw.QRadioButton("Pair each of them with a full boarder")
        self.put_them_in_pairs.clicked.connect(self.put_them_in_pairs_clicked)
        self.do_nothing = qtw.QRadioButton("Do nothing to the weekly boarders")
        self.do_nothing.clicked.connect(self.do_nothing_clicked)
        self.weekly_settings.addWidget(self.weekly_boarder_label)
        self.weekly_settings.addWidget(self.all_in_same_room)
        self.weekly_settings.addWidget(self.put_them_in_pairs)
        self.weekly_settings.addWidget(self.do_nothing)
        self.vlayout.addLayout(self.weekly_settings)

        self.time_warning_label = qtw.QLabel("Some certain combination of rooms could be extremely slow due to the amount of combinations generated. Would you like to speed up the calculation process? This may not produce the best results.")
        self.time_warning_check_box = qtw.QCheckBox("Speed up the process.")



        self.unwanted_pair_label = qtw.QLabel("Do you have any pairs of students you want to avoid putting into the same room?")    
        self.unwanted_pair_person_one = qtw.QComboBox()
        self.unwanted_pair_person_two = qtw.QComboBox()
        self.unwanted_pair_person_one.addItem("None")
        self.unwanted_pair_person_two.addItem("None")
        for student in self._students:
            self.unwanted_pair_person_one.addItem(student.name())
            self.unwanted_pair_person_two.addItem(student.name())
        self.add_unwanted_pair = qtw.QPushButton("+")
        self.add_unwanted_pair.clicked.connect(self.unwanted_add_clicked)
        self.unwanted_delete = qtw.QPushButton("-")
        self.unwanted_pair_display = qtw.QLabel("")
        self.unwanted_delete.clicked.connect(self.unwanted_delete_clicked)
        self.unwanted_pairs_buttons.addWidget(self.unwanted_delete)
        self.unwanted_pairs_buttons.addWidget(self.add_unwanted_pair)
        self.unwanted_pairs_hbox_layout.addWidget(self.unwanted_pair_person_one)
        self.unwanted_pairs_hbox_layout.addWidget(self.unwanted_pair_person_two)
        self.unwanted_pairs_vbox_layout.addWidget(self.unwanted_pair_label)
        self.unwanted_pairs_vbox_layout.addLayout(self.unwanted_pairs_hbox_layout)
        self.unwanted_pairs_vbox_layout.addLayout(self.unwanted_pairs_buttons)
        self.vlayout.addLayout(self.unwanted_pairs_vbox_layout)

        self.wanted_pair_label = qtw.QLabel("Do you have any pairs of students you want to put into the same room?")    
        self.wanted_pair_person_one = qtw.QComboBox()
        self.wanted_pair_person_two = qtw.QComboBox()
        self.wanted_pair_person_one.addItem("None")
        self.wanted_pair_person_two.addItem("None")
        for student in self._students:
            self.wanted_pair_person_one.addItem(student.name())
            self.wanted_pair_person_two.addItem(student.name())
        self.add_wanted_pair = qtw.QPushButton("+")
        self.add_wanted_pair.clicked.connect(self.wanted_add_clicked)
        self.wanted_delete = qtw.QPushButton("-")
        self.wanted_delete.clicked.connect(self.wanted_delete_clicked)
        self.wanted_pair_display = qtw.QLabel("")
        self.wanted_pairs_buttons.addWidget(self.wanted_delete)
        self.wanted_pairs_buttons.addWidget(self.add_wanted_pair)
        self.wanted_pairs_hbox_layout.addWidget(self.wanted_pair_person_one)
        self.wanted_pairs_hbox_layout.addWidget(self.wanted_pair_person_two)
        self.wanted_pairs_vbox_layout.addWidget(self.wanted_pair_label)
        self.wanted_pairs_vbox_layout.addLayout(self.wanted_pairs_hbox_layout)
        self.wanted_pairs_vbox_layout.addLayout(self.wanted_pairs_buttons)
        self.vlayout.addLayout(self.wanted_pairs_vbox_layout)

        self.pairs_display.addWidget(self.unwanted_pair_display)
        self.pairs_display.addWidget(self.wanted_pair_display)
        self.vlayout.addLayout(self.pairs_display)


        self.submit_button = qtw.QPushButton("Submit")
        self.submit_button.clicked.connect(self.submit_clicked)
        self.vlayout.addWidget(self.submit_button)


        self.setLayout(self.vlayout)
    
    def unwanted_add_clicked(self):
        if self.unwanted_pair_person_one.currentText() != "None" and self.unwanted_pair_person_two.currentText() != "None" and (self.unwanted_pair_person_one.currentText() != self.unwanted_pair_person_two.currentText()):
            self._unwanted_pairs.append([self.unwanted_pair_person_one.currentText(),self.unwanted_pair_person_two.currentText()])
            self._unwanted_combination_text = "The current unwanted pair(s) is/are:"
            for unwanted_pair in self._unwanted_pairs:
                if (f'{unwanted_pair[0]} and {unwanted_pair[1]}') in self._unwanted_combination_text or (f'{unwanted_pair[1]} and {unwanted_pair[0]}') in self._unwanted_combination_text:
                    self._unwanted_pairs.pop()
                else:
                    self._unwanted_combination_text += f'\n {unwanted_pair[0]} and {unwanted_pair[1]}'
        self.unwanted_pair_display.setText(self._unwanted_combination_text)

    def unwanted_delete_clicked(self):
        if len(self._unwanted_pairs) > 0:
            self._unwanted_pairs.pop()
            self._unwanted_combination_text = "The current unwanted pair(s) is/are:"
            if len(self._unwanted_pairs) == 0:
                self._unwanted_combination_text = ""
        for unwanted_pair in self._unwanted_pairs:
            self._unwanted_combination_text += f'\n {unwanted_pair[0]} and {unwanted_pair[1]}'
        self.unwanted_pair_display.setText(self._unwanted_combination_text)

    def wanted_add_clicked(self):
        if self.wanted_pair_person_one.currentText() != "None" and self.wanted_pair_person_two.currentText() != "None" and (self.wanted_pair_person_one.currentText() != self.wanted_pair_person_two.currentText()):
            self._wanted_pairs.append([self.wanted_pair_person_one.currentText(),self.wanted_pair_person_two.currentText()])
            self._wanted_combination_text = "The current wanted pair(s) is/are:"
            for wanted_pair in self._wanted_pairs:
                if (f'{wanted_pair[0]} and {wanted_pair[1]}') in self._wanted_combination_text or (f'{wanted_pair[1]} and {wanted_pair[0]}') in self._wanted_combination_text:
                    self._wanted_pairs.pop()
                else:
                    self._wanted_combination_text += f'\n {wanted_pair[0]} and {wanted_pair[1]}'
        self.wanted_pair_display.setText(self._wanted_combination_text)

    def wanted_delete_clicked(self):
        if len(self._wanted_pairs) > 0:
            self._wanted_pairs.pop()
            self._wanted_combination_text = "The current wanted pair(s) is/are:"
            if len(self._wanted_pairs) == 0:
                self._wanted_combination_text = ""
        for wanted_pair in self._wanted_pairs:
            self._wanted_combination_text += f'\n {wanted_pair[0]} and {wanted_pair[1]}'
        self.wanted_pair_display.setText(self._wanted_combination_text)
    
    def all_in_same_room_button_clicked(self):
        self._weekly_setting_text = "Put all the weekly boarders into the same room"
    
    def put_them_in_pairs_clicked(self):
        self._weekly_setting_text = "Pair each of them with a full boarder"

    def do_nothing_clicked(self):
        self._weekly_setting_text = "Nothing"

    def submit_clicked(self):
        no_of_people_due_to_assign_room = int(self.amount_of_one_man.currentText()) + (int(self.amount_of_two_man.currentText())*2) + (int(self.amount_of_three_man.currentText())*3)
        if no_of_people_due_to_assign_room != len(self._students):
            qtw.QMessageBox.warning(self, "incorrect number of rooms", "The amount of rooms assigned doesn't match the amount of students. Please double check.", qtw.QMessageBox.StandardButton.Ok)
        elif self.amount_of_combinations.text() == "":
            qtw.QMessageBox.warning(self, "empty combination", "Please enter the amount of combination needed.", qtw.QMessageBox.StandardButton.Ok)

        else:
            if self.result_window == None:
                amount_of_weekly = 0
                for student in self._students:
                    if student.is_weekly():
                        amount_of_weekly += 1
                if amount_of_weekly > 3 and self._weekly_setting_text == "Put all the weekly boarders into the same room":
                    qtw.QMessageBox.warning(self, "incorrect number of rooms", "There are more than three weekly boarders in this year group. Please choose another option.", qtw.QMessageBox.StandardButton.Ok)
                else:
                    self.close()
                    self.result_window = ResultWindow(self._students, self._unwanted_pairs, self._wanted_pairs, self._weekly_setting_text, int(self.amount_of_one_man.currentText()), int(self.amount_of_two_man.currentText()), int(self.amount_of_three_man.currentText()), int(self.amount_of_combinations.text()))
                    self.result_window.show()
            else:
                self.result_window.close()
                self.result_window = None

class ResultWindow(qtw.QWidget):
    def __init__(self, students, unwanted_pairs, wanted_pairs, weekly_settings, amount_of_one_man_rooms, amount_of_two_man_rooms, amount_of_three_man_rooms, amount_of_combinations):
        super().__init__()
        self.setWindowTitle("Results")
        self.result_layout = qtw.QVBoxLayout()
        self._students = students
        self._unwanted_pairs = unwanted_pairs
        self._wanted_pairs = wanted_pairs
        self._weekly_settings = weekly_settings
        self._amount_of_one_man_rooms = amount_of_one_man_rooms
        self._amount_of_two_man_rooms = amount_of_two_man_rooms
        self._amount_of_three_man_rooms = amount_of_three_man_rooms
        self._amount_of_combinations = amount_of_combinations
        self._rooming = rooming(self._students, self._unwanted_pairs, self._wanted_pairs, self._weekly_settings, self._amount_of_combinations)
        self._rooming.produce_roomings(self._amount_of_one_man_rooms,self._amount_of_two_man_rooms,self._amount_of_three_man_rooms)
        self._rooming.give_score_to_combinations()
        label = ""
        combination_no = 1
        for key in self._rooming.return_scores().keys():
            room_no = 1
            label += f"combination {combination_no}:\n \n"
            combination_no += 1
            for room in ast.literal_eval(key):
                label += f'Room {room_no}: '
                room_no += 1
                for student in room:
                    label += f'{student}, '
                label = label[:-2]
                label += "     "
            label += f'Score: {self._rooming.return_scores()[key]}'
            label += "\n"
            label += "\n"
        self.room_label = qtw.QLabel(label)
        print(label)
        self.result_layout.addWidget(self.room_label)
        self.setLayout(self.result_layout)



app = qtw.QApplication([])

window = MainWindow()
window.show()

app.exec()
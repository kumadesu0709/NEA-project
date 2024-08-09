from openpyxl.workbook import Workbook
from openpyxl import load_workbook


#load existing spreadsheet
wb = load_workbook("/Users/jamesguan/Desktop/CS A level NEA/NEA-project/NEA test sheet.xlsx")
#create the actual spreadsheet
ws = wb.active

print(ws["B100"].value)
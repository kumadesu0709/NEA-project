import random
import PyQt6.QtWidgets as qtw
from PyQt6.QtCore import Qt
from PyQt6.QtGui import * 
from openpyxl import load_workbook
import ast
from numbers_parser import Document


def generate_all_combination(students, max_no_of_people):
    result = []
    def backtrack(start, comb):
        if len(comb) == max_no_of_people:
            result.append(comb.copy())
        for i in range(start, len(students)):
            comb.append(students[i])
            backtrack(i+1, comb)
            comb.pop()
    backtrack(0, [])
    return result

class Student:

    def __init__(self, name: str, preferred_one: str, preferred_two: str, 
                 preferred_three: str, hated: str, is_weekly: bool):
        self._name = name
        self._preferred_one = preferred_one
        self._preferred_two = preferred_two
        self._preferred_three = preferred_three
        self._hated = hated
        if is_weekly == "Yes":
            self._is_weekly = True
        else:
            self._is_weekly = False

    def name(self):
        return self._name
    
    def preferred_one(self):
        return self._preferred_one
    
    def preferred_two(self):
        return self._preferred_two
    
    def preferred_three(self):
        return self._preferred_three
    
    def hated(self):
        return self._hated
    
    def is_weekly(self):
        return self._is_weekly
    

class Rooming:

    def __init__(self, student_in_yeargroup:list, unwanted_pairs:list, wanted_pairs:list, 
                 weekly_boarder_setting:str, amount_of_combinations: int):
        self._student_in_yeargroup = student_in_yeargroup
        self._unwanted_pairs = unwanted_pairs
        self._wanted_pairs = wanted_pairs
        self._amount_of_combinations = amount_of_combinations
        self._rooming_combinations = []
        self._rooming_scores = {}

        self._weekly_boarders = []
        for student in student_in_yeargroup:
            if student.is_weekly() == True:
                self._weekly_boarders.append(student.name())

        self._weekly_boarders_setting = weekly_boarder_setting
        if self._weekly_boarders_setting == "Put all the weekly boarders into the same room":
            if len(self._weekly_boarders) <= 3:
                self._wanted_pairs.append([self._weekly_boarders])
        elif self._weekly_boarders_setting == "Pair each of them with a full boarder":
            for i in range (len(self._weekly_boarders)):
                for j in range (i+1, len(self._weekly_boarders)):
                    self._unwanted_pairs.append([self._weekly_boarders[i],self._weekly_boarders[j]])


    def _check_existed_pupil(self,room_one, room_two):
        for student_one in room_one:
            if student_one in room_two:
                return True
        return False

    def pick_rooms(self, rooms:list, no_of_rooms:int):
        result = []
        def backtrack(start, comb):
            if len(comb) == no_of_rooms:
                result.append(comb.copy())
                return
            for i in range(start, len(rooms)):
                have_existed = False
                for existed_room in comb:
                    if self._check_existed_pupil(rooms[i], existed_room):
                        have_existed = True
                        break
                if have_existed == False:
                    comb.append(rooms[i])
                    backtrack(i+1,comb)
                    comb.pop()
        backtrack(0,[])
        return result
    
    def _produce_rooming_without_a_type_of_room(self, combs_room_one, combs_room_two):
        for comb_room_one in combs_room_one:
            for comb_room_two in combs_room_two:
                for room_one in comb_room_one:
                    for room_two in comb_room_two:
                        if self._check_existed_pupil(room_one, room_two):
                            continue
                        self._rooming_combinations.append([comb_room_one, comb_room_two])

    
    def produce_roomings(self, no_of_one_man_room:int,no_of_two_man_room:int,no_of_three_man_room:int):

        one_man_rooms = generate_all_combination(self._student_in_yeargroup,1)
        two_man_rooms = generate_all_combination(self._student_in_yeargroup,2)
        three_man_rooms = generate_all_combination(self._student_in_yeargroup,3)
        
        file = open('has_been_roommate.txt')
        contents = file.readlines() 
        file.close()
        file = open('has_been_roommate.txt', 'a')
        for pair in two_man_rooms:
            not_in_document = True
            for content in contents:
                if str([pair[0].name(), pair[1].name()]) in str(content) or str([pair[1].name(), pair[0].name()]) in str(content):
                    not_in_document = False
                    break
            if not_in_document == True:
                file.write(f'{[pair[0].name(), pair[1].name()]}0\n')
        file.close()
        file = open('has_been_roommate.txt')
        contents = file.readlines()

        combs_one_man_rooms = self.pick_rooms(one_man_rooms,no_of_one_man_room)
        combs_two_man_rooms = self.pick_rooms(two_man_rooms, no_of_two_man_room)
        combs_three_man_rooms = self.pick_rooms(three_man_rooms, no_of_three_man_room)

        if no_of_two_man_room == 0 and no_of_three_man_room == 0 and no_of_one_man_room > 0:
            for comb in combs_one_man_rooms:
                self._rooming_combinations.append([comb])
        
        elif no_of_one_man_room == 0 and no_of_three_man_room == 0 and no_of_two_man_room > 0:
            for comb in combs_two_man_rooms:
                self._rooming_combinations.append([comb])
        
        elif no_of_two_man_room == 0 and no_of_one_man_room == 0 and no_of_three_man_room > 0:
            for comb in combs_three_man_rooms:
                self._rooming_combinations.append([comb])

        elif no_of_one_man_room == 0 and no_of_two_man_room > 0 and no_of_three_man_room > 0:
            self._produce_rooming_without_a_type_of_room(combs_two_man_rooms, combs_three_man_rooms)

        elif no_of_one_man_room > 0 and no_of_two_man_room == 0 and no_of_three_man_room > 0:
            self._produce_rooming_without_a_type_of_room(combs_one_man_rooms, combs_three_man_rooms)

        elif no_of_one_man_room > 0 and no_of_two_man_room > 0 and no_of_three_man_room == 0:
            self._produce_rooming_without_a_type_of_room(combs_one_man_rooms, combs_two_man_rooms)

        elif no_of_one_man_room > 0 and no_of_two_man_room > 0 and no_of_three_man_room > 0:
            for comb_one_man_room in combs_one_man_rooms:
                for comb_two_man_room in combs_two_man_rooms:
                    for one_man_room in comb_one_man_room:
                        for two_man_room in comb_two_man_room:
                            if self._check_existed_pupil(one_man_room, two_man_room):
                                continue
                            for comb_three_man_room in combs_three_man_rooms:
                                for three_man_room in comb_three_man_room:
                                    if self._check_existed_pupil(three_man_room, one_man_room) or self._check_existed_pupil(two_man_room, three_man_room):
                                        continue
                                    self._rooming_combinations.append([comb_one_man_room,comb_two_man_room,comb_three_man_room])

    
    def _randomly_pick_combination(self):
        if len(self._rooming_combinations) > 100000:
            self._rooming_combinations = random.sample(self._rooming_combinations, len(self._rooming_combinations)//100)
            
    def _give_score_to_room(self,room:list):

        score = 0

        if len(room) == 2:

            if room[0].preferred_one() == room[1].name():
                score = score + 5
            if room[0].preferred_two() == room[1].name():
                score = score + 3
            if room[0].preferred_three() == room[1].name():
                score = score + 1
            if room[0].hated() == room[1]:
                score = score - 1


            if room[1].preferred_one() == room[0].name():
                score = score + 5
            if room[1].preferred_two() == room[0].name():
                score = score + 3
            if room[1].preferred_three() == room[0].name():
                score = score + 1
            if room[1].hated() == room[0].name():
                score = score - 1
        
        if len(room) == 3:

            if room[0].preferred_one() == room[1].name() or room[0].preferred_one == room[2].name():
                score = score + 5
            if room[0].preferred_two() == room[1].name() or room[0].preferred_two() == room[2].name():
                score = score + 3
            if room[0].preferred_three() == room[1].name() or room[0].preferred_three() == room[2].name():
                score = score + 1
            if room[0].hated() == room[1].name() or room[0].hated() == room[2].name():
                score = score - 1


            if room[1].preferred_one() == room[0].name() or room[1].preferred_one() == room[2].name():
                score = score + 5
            if room[1].preferred_two() == room[0].name() or room[1].preferred_two() == room[2].name():
                score = score + 3
            if room[1].preferred_three() == room[0].name() or room[1].preferred_three() == room[2].name():
                score = score + 1
            if room[1].hated() == room[0].name() or room[1].hated() == room[2].name():
                score = score - 1
            
            if room[2].preferred_one() == room[0].name() or room[2].preferred_one() == room[1].name():
                score = score + 5
            if room[2].preferred_two() == room[0].name() or room[2].preferred_two() == room[1].name():
                score = score + 3
            if room[2].preferred_three() == room[0].name() or room[2].preferred_three() == room[1].name():
                score = score + 1
            if room[2].hated() == room[0].name() or room[2].hated() == room[1].name():
                score = score - 1
            
        return score
    
    def _clean(self):
        rooming_combinations = self._rooming_combinations.copy()
                        
        if len(self._wanted_pairs) > 0:
            for combination in rooming_combinations:
                need_to_remove = True
                for rooms in combination:
                    for room in rooms:
                        print(room)
                        for wanted_pair in self._wanted_pairs:
                            if len(room) == 2:
                                    if set(wanted_pair).issubset(set([room[0].name(),room[1].name()])) == True:
                                        need_to_remove = False
                            if len(room) == 3:
                                if set(wanted_pair).issubset(set([room[0].name(),room[1].name(),room[2].name()])) == True:
                                    need_to_remove = False
                if need_to_remove == True:
                    if combination in self._rooming_combinations:
                        self._rooming_combinations.remove(combination)
        
        if len(self._unwanted_pairs) > 0:
            for combination in rooming_combinations:
                need_to_remove = False
                for rooms in combination:
                    for room in rooms:
                        for unwanted_pair in self._unwanted_pairs:
                            if len(room) == 2:
                                    if set(unwanted_pair).issubset(set([room[0].name(),room[1].name()])) == True:
                                        need_to_remove = True
                            if len(room) == 3:
                                if set(unwanted_pair).issubset(set([room[0].name(),room[1].name(),room[2].name()])) == True:
                                    need_to_remove = True
                if need_to_remove == True:
                    if combination in self._rooming_combinations:
                        self._rooming_combinations.remove(combination)

    def give_score_to_combinations(self,randomly_pick_room_on:bool):
        self._clean()
        if randomly_pick_room_on == True:
            self._randomly_pick_combination()
        for combination in self._rooming_combinations:
            score = 0
            rooms_in_combination = []
            for x_man_room_combination in combination:
                for room in x_man_room_combination:
                    score += self._give_score_to_room(room)
                    student_in_room = []
                    for student in room:
                        student_in_room.append(student.name())
                    rooms_in_combination.append(student_in_room)
            self._rooming_scores[str(rooms_in_combination)] = score
        self._rooming_scores = dict(sorted(self._rooming_scores.items(), key=lambda item: item[1], reverse=True))
        self._rooming_scores = {k: self._rooming_scores[k] for k in list(self._rooming_scores)[:self._amount_of_combinations]}

    def return_rooming(self):
        return self._rooming_combinations
    
    def return_scores(self):
        return self._rooming_scores

class MainWindow(qtw.QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Rooming Generator")
        self.setting_window = None

        layout = qtw.QGridLayout()
        vbox_layout = qtw.QVBoxLayout()

        form_layout = qtw.QFormLayout()
        self.directory = qtw.QLineEdit("")
        form_layout.addRow("Please type the directory of the excel file here (instructions below):", self.directory)
        vbox_layout.addLayout(form_layout)

        non_editable_line_edit = qtw.QLineEdit(parent=self)
        non_editable_line_edit.setReadOnly(True)
        non_editable_line_edit.setText("https://www.digitaltrends.com/computing/how-to-find-and-copy-a-file-path-on-mac/")
        vbox_layout.addWidget(non_editable_line_edit)

        self.submit_button = qtw.QPushButton("Submit")
        self.submit_button.clicked.connect(self.submit_button_clicked)
        vbox_layout.addWidget(self.submit_button)

        layout.addLayout(vbox_layout,0,0,1,3)

        widget = qtw.QWidget()
        widget.setLayout(layout)

        self.setCentralWidget(widget)
    
    def submit_button_clicked(self):
        if self.setting_window == None:

            window.close()
            wb = load_workbook(self.directory.text())
            ws = wb.active

            student_one = Student(ws["B2"].value, ws["C2"].value, ws["D2"].value, ws["E2"].value, ws["F2"].value, ws["G2"].value)
            student_two = Student(ws["B3"].value, ws["C3"].value, ws["D3"].value, ws["E3"].value, ws["F3"].value, ws["G3"].value)
            student_three = Student(ws["B4"].value, ws["C4"].value, ws["D4"].value, ws["E4"].value, ws["F4"].value, ws["G4"].value)
            student_four = Student(ws["B5"].value, ws["C5"].value, ws["D5"].value, ws["E5"].value, ws["F5"].value, ws["G5"].value)
            student_five = Student(ws["B6"].value, ws["C6"].value, ws["D6"].value, ws["E6"].value, ws["F6"].value, ws["G6"].value)
            student_six = Student(ws["B7"].value, ws["C7"].value, ws["D7"].value, ws["E7"].value, ws["F7"].value, ws["G7"].value)
            student_seven = Student(ws["B8"].value, ws["C8"].value, ws["D8"].value, ws["E8"].value, ws["F8"].value, ws["G8"].value)
            student_eight = Student(ws["B9"].value, ws["C9"].value, ws["D9"].value, ws["E9"].value, ws["F9"].value, ws["G9"].value)
            student_nine = Student(ws["B10"].value, ws["C10"].value, ws["D10"].value, ws["E10"].value, ws["F10"].value, ws["G10"].value)
            student_ten = Student(ws["B11"].value, ws["C11"].value, ws["D11"].value, ws["E11"].value, ws["F11"].value, ws["G11"].value)
            student_eleven = Student(ws["B12"].value, ws["C12"].value, ws["D12"].value, ws["E12"].value, ws["F12"].value, ws["G12"].value)
            student_twelve= Student(ws["B13"].value, ws["C13"].value, ws["D13"].value, ws["E13"].value, ws["F13"].value, ws["G13"].value)
            student_thirteen = Student(ws["B14"].value, ws["C14"].value, ws["D14"].value, ws["E14"].value, ws["F14"].value, ws["G14"].value)

            students = [student_one,student_two, student_three, student_four, 
                        student_five, student_six, student_seven, student_eight, 
                        student_nine, student_ten, student_eleven, student_twelve, 
                        student_thirteen]
            while (students[-1]).name() == None:
                students.pop()
            self.setting_window = SettingsWindow(students)
            self.setting_window.show()
        
        else:

            self.setting_window.close()
            self.setting_window = None


class SettingsWindow(qtw.QWidget):
    def __init__(self, students):

        super().__init__()
        self.setWindowTitle("Settings")

        self.unwanted_pairs = []
        self.wanted_pairs = []
        self.weekly_setting_text = ""
        self.use_previous_settings = False
        
        self.setting_window_layout = qtw.QVBoxLayout()
        self.unwanted_pairs_hbox_layout = qtw.QHBoxLayout()
        self.unwanted_pairs_vbox_layout = qtw.QVBoxLayout()
        self.wanted_pairs_hbox_layout = qtw.QHBoxLayout()
        self.wanted_pairs_vbox_layout = qtw.QVBoxLayout()
        self.unwanted_pairs_buttons = qtw.QHBoxLayout()
        self.wanted_pairs_buttons = qtw.QHBoxLayout()
        self.pairs_display = qtw.QHBoxLayout()

        self.amount_of_combination_layout = qtw.QHBoxLayout()
        self.weekly_settings = qtw.QVBoxLayout()
        self.room_settings = qtw.QGridLayout()


        self.students = students
        self.result_window = None
        self.show_previous_settings_window = None
        self._unwanted_combination_text = ""
        self._wanted_combination_text = ""

        self.one_man_label = qtw.QLabel("One Man Rooms")
        self.amount_of_one_man = qtw.QComboBox()
        self.amount_of_one_man.addItems(str(i) for i in range(0,(len(self.students)+1)))
        self.two_man_label = qtw.QLabel("Two Man Rooms")
        self.amount_of_two_man = qtw.QComboBox()
        self.amount_of_two_man.addItems(str(i) for i in range(0,(len(self.students)//2+1)))
        self.three_man_label = qtw.QLabel("Three Man Rooms")
        self.amount_of_three_man = qtw.QComboBox()
        self.amount_of_three_man.addItems(str(i) for i in range(0,(len(self.students)//3+1)))
        self.room_settings.addWidget(self.one_man_label,0,0)
        self.room_settings.addWidget(self.two_man_label,0,1)
        self.room_settings.addWidget(self.three_man_label,0,2)
        self.room_settings.addWidget(self.amount_of_one_man,1,0)
        self.room_settings.addWidget(self.amount_of_two_man,1,1)
        self.room_settings.addWidget(self.amount_of_three_man,1,2)
        self.setting_window_layout.addLayout(self.room_settings)


        self.amount_of_combinations_label = qtw.QLabel("How many combinations would you like to see?")
        self.amount_of_combinations_text = qtw.QLineEdit("")
        self.amount_of_combination_layout.addWidget(self.amount_of_combinations_label)
        self.amount_of_combination_layout.addWidget(self.amount_of_combinations_text)
        self.setting_window_layout.addLayout(self.amount_of_combination_layout)


        self.weekly_boarder_label = qtw.QLabel("Would you like to...")
        self.all_in_same_room = qtw.QRadioButton("Put all the weekly boarders into the same room")
        self.all_in_same_room.clicked.connect(self.all_in_same_room_button_clicked)
        self.put_them_in_pairs = qtw.QRadioButton("Pair each of them with a full boarder")
        self.put_them_in_pairs.clicked.connect(self.put_them_in_pairs_clicked)
        self.do_nothing = qtw.QRadioButton("Do nothing to the weekly boarders")
        self.do_nothing.clicked.connect(self.do_nothing_clicked)
        self.weekly_settings.addWidget(self.weekly_boarder_label)
        self.weekly_settings.addWidget(self.all_in_same_room)
        self.weekly_settings.addWidget(self.put_them_in_pairs)
        self.weekly_settings.addWidget(self.do_nothing)
        self.setting_window_layout.addLayout(self.weekly_settings)

        self.time_warning_check_box = qtw.QCheckBox("Would you like to speed up the calculation process? This may not produce the best results.")
        self.setting_window_layout.addWidget(self.time_warning_check_box)
        
        self.check_if_show_not_been_roommate = qtw.QCheckBox("Would you like to see the students that haven't been roommate before?")
        self.setting_window_layout.addWidget(self.check_if_show_not_been_roommate)
        
        self.unwanted_pair_label = qtw.QLabel("Do you have any pairs of students you want to avoid putting into the same room?")    
        self.unwanted_pair_person_one = qtw.QComboBox()
        self.unwanted_pair_person_two = qtw.QComboBox()
        self.unwanted_pair_person_one.addItem("None")
        self.unwanted_pair_person_two.addItem("None")
        for student in self.students:
            self.unwanted_pair_person_one.addItem(student.name())
            self.unwanted_pair_person_two.addItem(student.name())
        self.add_unwanted_pair = qtw.QPushButton("+")
        self.add_unwanted_pair.clicked.connect(self.unwanted_add_clicked)
        self.unwanted_delete = qtw.QPushButton("-")
        self.unwanted_pair_display = qtw.QLabel("")
        self.unwanted_delete.clicked.connect(self.unwanted_delete_clicked)
        self.unwanted_pairs_buttons.addWidget(self.unwanted_delete)
        self.unwanted_pairs_buttons.addWidget(self.add_unwanted_pair)
        self.unwanted_pairs_hbox_layout.addWidget(self.unwanted_pair_person_one)
        self.unwanted_pairs_hbox_layout.addWidget(self.unwanted_pair_person_two)
        self.unwanted_pairs_vbox_layout.addWidget(self.unwanted_pair_label)
        self.unwanted_pairs_vbox_layout.addLayout(self.unwanted_pairs_hbox_layout)
        self.unwanted_pairs_vbox_layout.addLayout(self.unwanted_pairs_buttons)
        self.setting_window_layout.addLayout(self.unwanted_pairs_vbox_layout)

        self.wanted_pair_label = qtw.QLabel("Do you have any pairs of students you want to put into the same room?")    
        self.wanted_pair_person_one = qtw.QComboBox()
        self.wanted_pair_person_two = qtw.QComboBox()
        self.wanted_pair_person_one.addItem("None")
        self.wanted_pair_person_two.addItem("None")
        for student in self.students:
            self.wanted_pair_person_one.addItem(student.name())
            self.wanted_pair_person_two.addItem(student.name())
        self.add_wanted_pair = qtw.QPushButton("+")
        self.add_wanted_pair.clicked.connect(self.wanted_add_clicked)
        self.wanted_delete = qtw.QPushButton("-")
        self.wanted_delete.clicked.connect(self.wanted_delete_clicked)
        self.wanted_pair_display = qtw.QLabel("")
        self.wanted_pairs_buttons.addWidget(self.wanted_delete)
        self.wanted_pairs_buttons.addWidget(self.add_wanted_pair)
        self.wanted_pairs_hbox_layout.addWidget(self.wanted_pair_person_one)
        self.wanted_pairs_hbox_layout.addWidget(self.wanted_pair_person_two)
        self.wanted_pairs_vbox_layout.addWidget(self.wanted_pair_label)
        self.wanted_pairs_vbox_layout.addLayout(self.wanted_pairs_hbox_layout)
        self.wanted_pairs_vbox_layout.addLayout(self.wanted_pairs_buttons)
        self.setting_window_layout.addLayout(self.wanted_pairs_vbox_layout)

        self.pairs_display.addWidget(self.unwanted_pair_display)
        self.pairs_display.addWidget(self.wanted_pair_display)
        self.setting_window_layout.addLayout(self.pairs_display)

        self.button_layout = qtw.QHBoxLayout()
        self.submit_button = qtw.QPushButton("Submit")
        self.submit_button.clicked.connect(self.submit_clicked)
        self.save_button = qtw.QPushButton("Save This Setting")
        self.save_button.clicked.connect(self.save_clicked)
        self.use_previous_settings_button = qtw.QPushButton('Use Previous Settings')
        self.use_previous_settings_button.clicked.connect(self.use_previous_settings_clicked)
        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.use_previous_settings_button)
        self.button_layout.addWidget(self.submit_button)
        
        self.setting_window_layout.addLayout(self.button_layout)
        self.setLayout(self.setting_window_layout)
    
    def unwanted_add_clicked(self):
        if self.unwanted_pair_person_one.currentText() != "None" and self.unwanted_pair_person_two.currentText() != "None" and (self.unwanted_pair_person_one.currentText() != self.unwanted_pair_person_two.currentText()):
            self.unwanted_pairs.append([self.unwanted_pair_person_one.currentText(),self.unwanted_pair_person_two.currentText()])
            self._unwanted_combination_text = "The current unwanted pair(s) is/are:"
            for unwanted_pair in self.unwanted_pairs:
                if (f'{unwanted_pair[0]} and {unwanted_pair[1]}') in self._unwanted_combination_text or (f'{unwanted_pair[1]} and {unwanted_pair[0]}') in self._unwanted_combination_text:
                    self.unwanted_pairs.pop()
                else:
                    self._unwanted_combination_text += f'\n {unwanted_pair[0]} and {unwanted_pair[1]}'
        self.unwanted_pair_display.setText(self._unwanted_combination_text)

    def unwanted_delete_clicked(self):
        if len(self.unwanted_pairs) > 0:
            self.unwanted_pairs.pop()
            self._unwanted_combination_text = "The current unwanted pair(s) is/are:"
            if len(self.unwanted_pairs) == 0:
                self._unwanted_combination_text = ""
        for unwanted_pair in self.unwanted_pairs:
            self._unwanted_combination_text += f'\n {unwanted_pair[0]} and {unwanted_pair[1]}'
        self.unwanted_pair_display.setText(self._unwanted_combination_text)

    def wanted_add_clicked(self):
        if self.wanted_pair_person_one.currentText() != "None" and self.wanted_pair_person_two.currentText() != "None" and (self.wanted_pair_person_one.currentText() != self.wanted_pair_person_two.currentText()):
            self.wanted_pairs.append([self.wanted_pair_person_one.currentText(),self.wanted_pair_person_two.currentText()])
            self._wanted_combination_text = "The current wanted pair(s) is/are:"
            for wanted_pair in self.wanted_pairs:
                if (f'{wanted_pair[0]} and {wanted_pair[1]}') in self._wanted_combination_text or (f'{wanted_pair[1]} and {wanted_pair[0]}') in self._wanted_combination_text:
                    self.wanted_pairs.pop()
                else:
                    self._wanted_combination_text += f'\n {wanted_pair[0]} and {wanted_pair[1]}'
        self.wanted_pair_display.setText(self._wanted_combination_text)

    def wanted_delete_clicked(self):
        if len(self.wanted_pairs) > 0:
            self.wanted_pairs.pop()
            self._wanted_combination_text = "The current wanted pair(s) is/are:"
            if len(self.wanted_pairs) == 0:
                self._wanted_combination_text = ""
        for wanted_pair in self.wanted_pairs:
            self._wanted_combination_text += f'\n {wanted_pair[0]} and {wanted_pair[1]}'
        self.wanted_pair_display.setText(self._wanted_combination_text)
    
    def all_in_same_room_button_clicked(self):
        self.weekly_setting_text = "Put all the weekly boarders into the same room"
    
    def put_them_in_pairs_clicked(self):
        self.weekly_setting_text = "Pair each of them with a full boarder"

    def do_nothing_clicked(self):
        self.weekly_setting_text = "Nothing"

    def save_clicked(self):
        with open ('room_settings.txt', 'w') as file:
            file.write(f'{self.unwanted_pairs}\n')
            file.write(f'{self.wanted_pairs}\n')
            file.write(f'{self.weekly_setting_text}\n')
            file.write(f'{self.amount_of_combinations_text.text()}\n')
            file.write(f'{self.amount_of_one_man.currentText()}\n')
            file.write(f'{self.amount_of_two_man.currentText()}\n')
            file.write(f'{self.amount_of_three_man.currentText()}\n')
            file.write(f'{self.time_warning_check_box.isChecked()}\n')
            file.write(f'{self.check_if_show_not_been_roommate.isChecked()}')
        
    def use_previous_settings_clicked(self):
        self.use_previous_settings = True
        
        if self.show_previous_settings_window == None:
            self.show_previous_settings_window = CheckIfUsePreviousSettingsWindow(self.students)
            self.show_previous_settings_window.show()
        else:
            self.show_previous_settings_window.close()
            self.show_previous_settings_window = None

    def submit_clicked(self):
        no_of_people_due_to_assign_room = int(self.amount_of_one_man.currentText()) + (int(self.amount_of_two_man.currentText())*2) + (int(self.amount_of_three_man.currentText())*3)
        if no_of_people_due_to_assign_room != len(self.students):
            qtw.QMessageBox.warning(self, "incorrect number of rooms", "The amount of rooms assigned doesn't match the amount of students. Please double check.", qtw.QMessageBox.StandardButton.Ok)
        elif self.amount_of_combinations_text.text() == "0":
            qtw.QMessageBox.warning(self, "empty combination", "Please enter the amount of combination needed.", qtw.QMessageBox.StandardButton.Ok)

        else:
            if self.result_window == None:
                amount_of_weekly = 0
                for student in self.students:
                    if student.is_weekly():
                        amount_of_weekly += 1
                if amount_of_weekly > 3 and self.weekly_setting_text == "Put all the weekly boarders into the same room":
                    qtw.QMessageBox.warning(self, "incorrect number of rooms", "There are more than three weekly boarders in this year group. Please choose another option.", qtw.QMessageBox.StandardButton.Ok)
                else:
                    self.result_window = ResultWindow(self.students, self.unwanted_pairs, self.wanted_pairs, 
                                                      self.weekly_setting_text, int(self.amount_of_combinations_text.text()),
                                                      int(self.amount_of_one_man.currentText()), int(self.amount_of_two_man.currentText()),
                                                      int(self.amount_of_three_man.currentText()), self.time_warning_check_box.isChecked(), 
                                                      self.check_if_show_not_been_roommate.isChecked())
                    self.result_window.show()
                    self.close()
            else:
                self.result_window.close()
                self.result_window = None

class CheckIfUsePreviousSettingsWindow(qtw.QWidget):
    
    def __init__(self, students):
        
        super().__init__()
        file = open('room_settings.txt')
        settings = file.readlines()
        file.close()
        self.result_window = None
        
        self.students = students
        self.show_previous_settings_layout = qtw.QVBoxLayout()
        self.show_previous_settings_button_layout = qtw.QHBoxLayout()
        self.ok_button = qtw.QPushButton("Ok")
        self.ok_button.clicked.connect(self.ok_clicked)
        self.cancel_button = qtw.QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.cancel_clicked)
        self.show_previous_settings_button_layout.addWidget(self.ok_button)
        self.show_previous_settings_button_layout.addWidget(self.cancel_button)
        
        setting_string = 'Are you sure you want to use the previous settings? The previous setting is:\n'
        setting_string += 'wanted pair(s): '
        for wanted_pair in ast.literal_eval(settings[1][:-1]):
            setting_string += f'{wanted_pair[0]} and {wanted_pair[1]}, '
        setting_string = f'{setting_string[:-2]}\n'
        setting_string += 'unwanted pair(s): '
        for unwanted_pair in ast.literal_eval(settings[0][:-1]):
            setting_string += f'{unwanted_pair[0]} and {unwanted_pair[1]}, '
        setting_string = f'{setting_string[:-2]}\n'
        setting_string += f'What to do with weekly boarders: {settings[2][:-1]}\n'
        setting_string += f'Generate {settings[3][:-1]} combinations\n'
        setting_string += f'Assign {settings[4][:-1]} one-man rooms, {settings[5][:-1]} two-man rooms, and {settings[6][:-1]} three-man rooms to this yeargroup\n'
        if ast.literal_eval(settings[7][:-1]) and ast.literal_eval(settings[8]):
            setting_string += "Speed up the calculation process\nShow the students that hasn't been roommate"
        elif ast.literal_eval(settings[7][:-1]) and not ast.literal_eval(settings[8]):
            setting_string += "Speed up the calculation process"
            
        self.settings_label = qtw.QLabel(setting_string)
        self.show_previous_settings_layout.addWidget(self.settings_label)
        self.show_previous_settings_layout.addLayout(self.show_previous_settings_button_layout)
        self.setLayout(self.show_previous_settings_layout)
        
    def cancel_clicked(self):
        self.close()
    
    def ok_clicked(self):
        
        if self.result_window == None:
            settings_file = open("room_settings.txt")
            settings = settings_file.readlines()
            settings_file.close()
            self.result_window = ResultWindow(self.students, ast.literal_eval(settings[0][:-1]), 
                                              ast.literal_eval(settings[1][:-1]), 
                                              settings[2][:-1], 
                                              ast.literal_eval(settings[3][:-1]), 
                                              ast.literal_eval(settings[4][:-1]), 
                                              ast.literal_eval(settings[5][:-1]),
                                              ast.literal_eval(settings[6][:-1]), 
                                              ast.literal_eval(settings[7][:-1]), 
                                              ast.literal_eval(settings[8]))
            self.result_window.show()
            self.close()
        else:
            self.result_window.close()
            self.result_window = None

class ResultWindow(qtw.QWidget):
    
    def __init__(self, students, unwanted_pairs, wanted_pairs, weekly_settings, amount_of_combinations,
                 amount_of_one_man, amount_of_two_man, amount_of_three_man, speed_up_calc, show_hasnt_been_roommate):
        
        super().__init__()
        
        self.students = students
        self.unwanted_pairs = unwanted_pairs
        self.wanted_pairs = wanted_pairs
        self.weekly_settings = weekly_settings
        self.amount_of_combinations = amount_of_combinations
        self.amount_of_one_man_rooms = amount_of_one_man
        self.amount_of_two_man_rooms = amount_of_two_man
        self.amount_of_three_man_rooms = amount_of_three_man
        self.speed_up_calc = speed_up_calc
        self.show_hasnt_been_roommate = show_hasnt_been_roommate
        
        self.setWindowTitle("Results")
        self.result_layout = qtw.QVBoxLayout()
        self.reminder_label = qtw.QLabel('Click the combination to select the combination, or click "Create Own Combination" to create your own combination')
        self.reminder_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.result_layout.addWidget(self.reminder_label)
        self._returned_rooming_list = []
        
        self.settings_window = None
        self.edit_window = None
        self.output_setting_window = None
        self._rooming = Rooming(self.students, self.unwanted_pairs, self.wanted_pairs, self.weekly_settings, self.amount_of_combinations)
        self._rooming.produce_roomings(self.amount_of_one_man_rooms,self.amount_of_two_man_rooms,self.amount_of_three_man_rooms)
        self._rooming.give_score_to_combinations(self.speed_up_calc)
        self._rooming_scores = self._rooming.return_scores()
        
        if self.show_hasnt_been_roommate:
            file = open('has_been_roommate.txt')
            contents = file.readlines()
            hasnt_been_roommate_string = ''
            for content in contents:
                if content[len(content)-2] == '0':
                    pair = ast.literal_eval(content[0:len(content)-2])
                    hasnt_been_roommate_string += f'{pair[0]} and {pair[1]}, '
            hasnt_been_roommate_string = hasnt_been_roommate_string[0:len(hasnt_been_roommate_string)-2]
            hasnt_been_roommate_string += " hasn't been roommate yet."
            self.hasnt_been_roommate_label = qtw.QLabel(hasnt_been_roommate_string)
            self.hasnt_been_roommate_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.result_layout.addWidget(self.hasnt_been_roommate_label)
        
        self.create_own_combination = qtw.QPushButton("Create Own Combination")
        self.create_own_combination.clicked.connect(self.create_own_combination_clicked)
        self.back_to_settings = qtw.QPushButton("Back to Settings")
        self.back_to_settings.clicked.connect(self.back_to_settings_clicked)
        
        self.result_page_buttons_layout = qtw.QHBoxLayout()
        
        if len(self._rooming_scores) > 10:
            self.next_page_button = qtw.QPushButton(">")
            self.back_page_button = qtw.QPushButton("<")
            self.next_page_button.clicked.connect(self.next_clicked)
            self.back_page_button.clicked.connect(self.back_clicked)
            self.result_page_buttons_layout.addWidget(self.back_page_button)
            self.result_page_buttons_layout.addWidget(self.create_own_combination)
            self.result_page_buttons_layout.addWidget(self.back_to_settings)
            self.result_page_buttons_layout.addWidget(self.next_page_button)
        else:
            self.result_page_buttons_layout.addWidget(self.create_own_combination)
            self.result_page_buttons_layout.addWidget(self.back_to_settings)
        
        self._combination_numbers = ["x"]*(self.amount_of_combinations+10)
        self._room_labels = ["x"]*(self.amount_of_combinations+10)
        self._ranking_rooms = {}
        
        for i in range (0, len(self._combination_numbers)):
            self._combination_numbers[i] = f'Combination {i+1}:'
        
        position = 0
        for key in (self._rooming_scores.keys()):
            label = ""
            room_no = 1
            for room in ast.literal_eval(key):
                label += f'Room {room_no}: '
                room_no += 1
                for student in room:
                    label += f'{student}, '
                label = label[:-2]
                label += "     "
            label += f'Score: {self._rooming.return_scores()[key]}'
            self._room_labels[position] = label
            position += 1
        
        for i in range(len(self._combination_numbers)-1):
            self._ranking_rooms[self._combination_numbers[i]] = self._room_labels[i]
        
        self._start_with = 0
        if len(self._rooming_scores) > 11:
            self._end_with = 10
        else:
            self._end_with = len(self._rooming_scores)
        
        self.combination_layout = qtw.QVBoxLayout()
        self.create_combination_layout()
        
        self.result_layout.addLayout(self.combination_layout)
        self.result_layout.addLayout(self.result_page_buttons_layout)
        self.setLayout(self.result_layout)

    def next_clicked(self):
        for i in reversed(range(self.combination_layout.count())): 
            self.combination_layout.itemAt(i).widget().setParent(None)
        if self._end_with != len(self._rooming_scores) and (self._start_with + 10) < len(self._rooming_scores)-1:
            self._start_with = self._start_with + 10
            
        if (self._end_with + 10) < len(self._rooming_scores):
            self._end_with = self._end_with + 10
        else:
            self._end_with = len(self._rooming_scores)-1
        self.create_combination_layout()
    
    def back_clicked(self):
        for i in reversed(range(self.combination_layout.count())): 
            self.combination_layout.itemAt(i).widget().setParent(None)
        if (self._start_with - 10) > 0:
            if (self._end_with % 10) == 0:
                self._start_with = self._start_with - 10
            else:
                self._start_with = self._start_with - 10
        else:
            self._start_with = 0
        if (self._start_with + 10) < len(self._rooming_scores)-1:
            self._end_with = self._start_with + 10
        else:
            self._end_with = self._start_with + (len(self._rooming_scores)%10)
        self.create_combination_layout()
    
    def create_combination_layout(self):
        
        self.rooms_and_combination_one = qtw.QPushButton(self._room_labels[self._start_with])
        self.rooms_and_combination_one.clicked.connect(self.combination_button_clicked)
        self.rooms_and_combination_two = qtw.QPushButton(self._room_labels[self._start_with +1])
        self.rooms_and_combination_two.clicked.connect(self.combination_button_clicked)
        self.rooms_and_combination_three = qtw.QPushButton(self._room_labels[self._start_with +2])
        self.rooms_and_combination_three.clicked.connect(self.combination_button_clicked)
        self.rooms_and_combination_four = qtw.QPushButton(self._room_labels[self._start_with + 3])
        self.rooms_and_combination_four.clicked.connect(self.combination_button_clicked)
        self.rooms_and_combination_five = qtw.QPushButton(self._room_labels[self._start_with +4])
        self.rooms_and_combination_five.clicked.connect(self.combination_button_clicked)
        self.rooms_and_combination_six = qtw.QPushButton(self._room_labels[self._start_with +5])
        self.rooms_and_combination_six.clicked.connect(self.combination_button_clicked)
        self.rooms_and_combination_seven = qtw.QPushButton(self._room_labels[self._start_with +6])
        self.rooms_and_combination_seven.clicked.connect(self.combination_button_clicked)
        self.rooms_and_combination_eight = qtw.QPushButton(self._room_labels[self._start_with +7])
        self.rooms_and_combination_eight.clicked.connect(self.combination_button_clicked)
        self.rooms_and_combination_nine = qtw.QPushButton(self._room_labels[self._start_with +8])
        self.rooms_and_combination_nine.clicked.connect(self.combination_button_clicked)
        self.rooms_and_combination_ten = qtw.QPushButton(self._room_labels[self._start_with + 9])
        self.rooms_and_combination_ten.clicked.connect(self.combination_button_clicked)
        
        rooms_and_combination_buttons = [self.rooms_and_combination_one, self.rooms_and_combination_two, self.rooms_and_combination_three, 
                                         self.rooms_and_combination_four, self.rooms_and_combination_five, self.rooms_and_combination_six, 
                                         self.rooms_and_combination_seven, self.rooms_and_combination_eight, self.rooms_and_combination_nine, 
                                         self.rooms_and_combination_ten]
        
        while rooms_and_combination_buttons[-1].text() == "x":
            rooms_and_combination_buttons.pop()
        
        start = self._start_with
        for button in rooms_and_combination_buttons:
            self.combination_layout.addWidget(qtw.QLabel(self._combination_numbers[start]))
            self.combination_layout.addWidget(button)
            start += 1
            
        self.result_layout.addLayout(self.combination_layout)
    
    def combination_button_clicked(self):
        button_clicked = self.sender()
        text = button_clicked.text()
        text = text.replace("     ", ",")
        text = text.replace(":", ",")
        text = text.replace(" ", "")
        room_label_list = list(text.split(","))[:-2]
        
        for i in range (self.amount_of_one_man_rooms):
            room_label = []
            room_label = room_label_list[0:2]
            for label in room_label:
                room_label_list.remove(label)
            self._returned_rooming_list.append(room_label)
        
        for i in range (self.amount_of_two_man_rooms):
            room_label = []
            room_label = room_label_list[0:3]
            for label in room_label:
                room_label_list.remove(label)
            self._returned_rooming_list.append(room_label)
        
        for i in range (self.amount_of_three_man_rooms):
            room_label = []
            room_label = room_label_list[0:4]
            for label in room_label:
                room_label_list.remove(label)
            self._returned_rooming_list.append(room_label)
        
        if self.output_setting_window == None:
            self.output_setting_window = OutputSettingsWindow(self._returned_rooming_list)
            self.output_setting_window.show()
        else:
            self.output_setting_window.close()
            self.output_setting_window = None
    
    def back_to_settings_clicked(self):
        
        if self.settings_window == None:
            self.close()
            self.settings_window = SettingsWindow(self.students)
            self.settings_window.show()
        else:
            self.settings_window.close()
            self.settings_window = None
    
    def create_own_combination_clicked(self):
        
        if self.edit_window == None:
            self.edit_window = EditWindow(self.students, self.amount_of_one_man_rooms, self.amount_of_two_man_rooms, self.amount_of_three_man_rooms)
            self.edit_window.show()
        else:
            self.edit_window.close()
            self.edit_window = None

class EditWindow(qtw.QWidget):
    def __init__ (self,students, amount_of_one_man, amount_of_two_man, amount_of_three_man):
        
        super().__init__()
        self.setWindowTitle("Create Own Combination")
        
        self.amount_of_one_man_rooms = amount_of_one_man
        self.amount_of_two_man_rooms = amount_of_two_man
        self.amount_of_three_man_rooms = amount_of_three_man
        self.students = students
        self._output_settings_window = None
        self._returned_room_list = []
        
        self.room_one_combo_box = qtw.QComboBox()
        self.room_two_combo_box = qtw.QComboBox()
        self.room_three_combo_box = qtw.QComboBox()
        self.room_four_combo_box = qtw.QComboBox()
        self.room_five_combo_box = qtw.QComboBox()
        self.room_six_combo_box = qtw.QComboBox()
        self.room_seven_combo_box = qtw.QComboBox()
        self.room_eight_combo_box = qtw.QComboBox()
        self.room_nine_combo_box = qtw.QComboBox()
        self.room_ten_combo_box = qtw.QComboBox()
        self.room_eleven_combo_box = qtw.QComboBox()
        self.room_twelve_combo_box = qtw.QComboBox()
        self.room_thirteen_combo_box = qtw.QComboBox()
        
        self._room_combo_boxes = [self.room_one_combo_box,self.room_two_combo_box,self.room_three_combo_box,
                                  self.room_four_combo_box,self.room_five_combo_box,self.room_six_combo_box,
                                  self.room_seven_combo_box,self.room_eight_combo_box,self.room_nine_combo_box,
                                  self.room_ten_combo_box, self.room_eleven_combo_box, self.room_twelve_combo_box, 
                                  self.room_thirteen_combo_box]
        
        self.combo_boxes_layout = qtw.QHBoxLayout()
        self.edit_layout = qtw.QVBoxLayout()
        
        for combo_box in self._room_combo_boxes:
            for student in self.students:
                combo_box.addItem(student.name())
        
        room_labels = []
        
        for i in range (1, (self.amount_of_one_man_rooms+self.amount_of_two_man_rooms+self.amount_of_three_man_rooms+1)):
            room_labels.append(qtw.QLabel(f"Room {i}"))
        for label in room_labels:
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self._returned_room_list.append([label.text()])
        
        self.combo_box_layout_one = qtw.QHBoxLayout()
        self.combo_box_layout_two = qtw.QHBoxLayout()
        self.combo_box_layout_three = qtw.QHBoxLayout()
        self.combo_box_layout_four = qtw.QHBoxLayout()
        self.combo_box_layout_five = qtw.QHBoxLayout()
        self.combo_box_layout_six = qtw.QHBoxLayout()
        self.combo_box_layout_seven = qtw.QHBoxLayout()
        self.combo_box_layout_eight = qtw.QHBoxLayout()
        self.combo_box_layout_nine = qtw.QHBoxLayout()
        self.combo_box_layout_ten = qtw.QHBoxLayout()
        self.combo_box_layout_eleven = qtw.QHBoxLayout()
        self.combo_box_layout_twelve = qtw.QHBoxLayout()
        self.combo_box_layout_thirteen = qtw.QHBoxLayout()
        
        combo_box_layouts = [self.combo_box_layout_one, self.combo_box_layout_two, self.combo_box_layout_three, 
                             self.combo_box_layout_four, self.combo_box_layout_five, self.combo_box_layout_six, 
                             self.combo_box_layout_seven, self.combo_box_layout_eight, self.combo_box_layout_nine, 
                             self.combo_box_layout_ten, self.combo_box_layout_eleven, self.combo_box_layout_twelve, 
                             self.combo_box_layout_thirteen]
        
        combo_box_layout_number = 0
        combo_box_number = 0
        
        for i in range(self.amount_of_one_man_rooms):
            
            combo_box_layouts[combo_box_layout_number].addWidget(self._room_combo_boxes[combo_box_number])
            combo_box_layout_number += 1
            combo_box_number += 1
        
        for i in range(self.amount_of_two_man_rooms):
            
            combo_box_layouts[combo_box_layout_number].addWidget(self._room_combo_boxes[combo_box_number])
            combo_box_number += 1
            combo_box_layouts[combo_box_layout_number].addWidget(self._room_combo_boxes[combo_box_number])
            combo_box_number += 1
            
            combo_box_layout_number += 1
        
        for i in range(self.amount_of_three_man_rooms):
            
            combo_box_layouts[combo_box_layout_number].addWidget(self._room_combo_boxes[combo_box_number])
            combo_box_number += 1
            combo_box_layouts[combo_box_layout_number].addWidget(self._room_combo_boxes[combo_box_number])
            combo_box_number += 1
            combo_box_layouts[combo_box_layout_number].addWidget(self._room_combo_boxes[combo_box_number])
            combo_box_number += 1
            
            combo_box_layout_number += 1
        
        self.room_layout_one = qtw.QVBoxLayout()
        self.room_layout_two = qtw.QVBoxLayout()
        self.room_layout_three = qtw.QVBoxLayout()
        self.room_layout_four = qtw.QVBoxLayout()
        self.room_layout_five = qtw.QVBoxLayout()
        self.room_layout_six = qtw.QVBoxLayout()
        self.room_layout_seven = qtw.QVBoxLayout()
        self.room_layout_eight = qtw.QVBoxLayout()
        self.room_layout_nine = qtw.QVBoxLayout()
        self.room_layout_ten = qtw.QVBoxLayout()
        self.room_layout_eleven = qtw.QVBoxLayout()
        self.room_layout_twelve = qtw.QVBoxLayout()
        self.room_layout_thirteen = qtw.QVBoxLayout()
        
        room_layouts = [self.room_layout_one, self.room_layout_two, self.room_layout_three, 
                        self.room_layout_four, self.room_layout_five, self.room_layout_six, 
                        self.room_layout_seven, self.room_layout_eight, self.room_layout_nine, 
                        self.room_layout_ten, self.room_layout_eleven, self.room_layout_twelve, 
                        self.room_layout_thirteen]
        room_layouts = room_layouts[0:(self.amount_of_one_man_rooms+self.amount_of_two_man_rooms+self.amount_of_three_man_rooms)]
        
        
        for i in range((self.amount_of_one_man_rooms+self.amount_of_two_man_rooms+self.amount_of_three_man_rooms)):
            room_layouts[i].addWidget(room_labels[i])
            room_layouts[i].addLayout(combo_box_layouts[i])
        
        for room_layout in room_layouts:
            self.combo_boxes_layout.addLayout(room_layout)

        self.edit_layout.addLayout(self.combo_boxes_layout)
        
        self.submit_combination_button = qtw.QPushButton("submit")
        self.submit_combination_button.clicked.connect(self.submit_button_clicked)
        self.edit_layout.addWidget(self.submit_combination_button)
        
        self.setLayout(self.edit_layout)
    
    def submit_button_clicked(self):
        used_students = []
        can_pass_on = True
        self._room_combo_boxes_copy = self._room_combo_boxes[0:(self.amount_of_one_man_rooms+self.amount_of_two_man_rooms+self.amount_of_three_man_rooms+1)]
        
        for combo_box in self._room_combo_boxes_copy:
            if combo_box.currentText() not in used_students:
                used_students.append(combo_box.currentText())
            else:
                qtw.QMessageBox.warning(self, "repeated student", "You've accidentally used the same student(s) twice. Please double check.", qtw.QMessageBox.StandardButton.Ok)
                can_pass_on = False
                break
        if can_pass_on == True:
            if self._output_settings_window == None:
                returned_room_list_index = 0
                combo_box_index = 0
                for i in range(self.amount_of_one_man_rooms):
                    self._returned_room_list[returned_room_list_index].append(self._room_combo_boxes[combo_box_index].currentText())
                    returned_room_list_index += 1
                    combo_box_index += 1
                for i in range(self.amount_of_two_man_rooms):
                    self._returned_room_list[returned_room_list_index].append(self._room_combo_boxes[combo_box_index].currentText())
                    combo_box_index += 1
                    self._returned_room_list[returned_room_list_index].append(self._room_combo_boxes[combo_box_index].currentText())
                    combo_box_index += 1
                    returned_room_list_index += 1
                for i in range(self.amount_of_three_man_rooms):
                    self._returned_room_list[returned_room_list_index].append(self._room_combo_boxes[combo_box_index].currentText())
                    combo_box_index += 1
                    self._returned_room_list[returned_room_list_index].append(self._room_combo_boxes[combo_box_index].currentText())
                    combo_box_index += 1
                    self._returned_room_list[returned_room_list_index].append(self._room_combo_boxes[combo_box_index].currentText())
                    combo_box_index += 1
                    returned_room_list_index += 1
                self._output_settings_window = OutputSettingsWindow(self._returned_room_list)
                self._output_settings_window.show()
            else:
                self._output_settings_window.close()
                self._output_settings_window = None


class OutputSettingsWindow(qtw.QWidget):
    
    def __init__(self, rooming_text_list):
        
        super(OutputSettingsWindow, self).__init__()
        self.has_been_imported = None
        
        file = open('has_been_roommate.txt')
        self._contents = file.readlines() 
        file.close()
        
        self.setWindowTitle("Output Settings")
        qtw.QApplication.closeAllWindows()
        self._rooming_text_list = rooming_text_list
        self.submit_button = qtw.QPushButton("Submit")
        self.submit_button.clicked.connect(self.submit_clicked)
        self.output_setting_layout = qtw.QVBoxLayout()
        form_layout_for_document = qtw.QFormLayout()
        self.document_name = qtw.QLineEdit()
        self.year_group_text= qtw.QLineEdit()
        self.check_if_create_new_number_sheet = qtw.QCheckBox("Would you like to create a new document?")
        form_layout_for_document.addRow('Please enter the name of the number sheet (or the name that you want to give to the sheet)', self.document_name)
        form_layout_for_document.addRow('Which year group is this for?', self.year_group_text)
        self.output_setting_layout.addLayout(form_layout_for_document)
        self.output_setting_layout.addWidget(self.check_if_create_new_number_sheet)
        self.output_setting_layout.addWidget(self.submit_button)
        self.setLayout(self.output_setting_layout)
        
    def submit_clicked(self):
        
        if self.check_if_create_new_number_sheet.isChecked():
            doc = Document()
        else:
            if self.document_name.text()[-8:] == ".numbers":
                doc = Document(self.document_name.text())
            else:
                doc = Document(f"{self.document_name.text()}.numbers")
        doc.add_sheet(self.year_group_text.text(), self.year_group_text.text())
        sheet = doc.sheets[self.year_group_text.text()]
        table = sheet.tables[self.year_group_text.text()]
        row = 1
        column = 0
        for room in self._rooming_text_list:
            column = 0
            for label in room:
                table.write(row, column, label)
                column += 1
            row+= 1
        if self.document_name.text()[-8:] == ".numbers":
            doc.save(self.document_name.text())
        else:
            doc.save(f'{self.document_name.text()}.numbers')

        copy_of_contents = self._contents.copy()
        
        for rooming_text_list in self._rooming_text_list:
            list_of_combination_in_room = generate_all_combination(rooming_text_list[1:len(rooming_text_list)], 2)
            for combination in list_of_combination_in_room:
                for i in range(len(self._contents)):
                    if str(combination) in self._contents[i] or str([combination[1],combination[0]]) in self._contents[i]:
                        copy_of_contents[i] = f'{str(combination)}1\n'
        
        with open('has_been_roommate.txt', 'w') as file:
            for content in copy_of_contents:
                file.write(content)
        file.close()
        
        if self.has_been_imported == None:
            self.close()
            self.has_been_imported = HaveExportedWindow()
            self.has_been_imported.show()
        else:
            self.has_been_imported.close()
            self.has_been_imported_window = None

class HaveExportedWindow(qtw.QWidget):
    
    def __init__(self):
        
        super().__init__()
        self.congrats_layout = qtw.QVBoxLayout()
        self.congrats_label = qtw.QLabel("Your choice has been exported onto the chosen number sheet")
        self.congrats_layout.addWidget(self.congrats_label)
        self.setLayout(self.congrats_layout)

app = qtw.QApplication([])

window = MainWindow()
window.show()

app.exec()